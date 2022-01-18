from matplotlib import pyplot                       # Импорт модуля pyplot из библиотеки matplotlib
from openpyxl import load_workbook                  # Импорт модуля load_workbook из модуля openpyxl

wb = load_workbook('data_analysis_lab.xlsx')        # Открываем таблицу и создаём ссылающуюся на не переменную
sheet = wb['Data']                                  # Создаём переменную ссылающуюся на лист "Data"

def getvalue(x):                                    # Функция возвращает содержимое "x"
    return x.value

Y = list(map(getvalue, sheet['A'][1:]))             # Создаём переменные со списками значений по столбцам
T = list(map(getvalue, sheet['B'][1:]))
D = list(map(getvalue, sheet['C'][1:]))
S = list(map(getvalue, sheet['D'][1:]))

pyplot.plot(Y, T, label="Температура")              # Загоняем списки для формирования графиков
pyplot.plot(Y, D, label="Относительная температура")
pyplot.plot(Y, S, label="Солнечная Активность")

pyplot.show()                                       # Отрисовываем график